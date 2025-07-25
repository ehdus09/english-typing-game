import random as r
import time as t
import csv
import openpyxl
import os
import sys


# 프로그램 소개
def intro():
    print("영컴타자연습에 오신걸 환영합니다.")
    t.sleep(1)
    print("영어단어를 타자연습을 하면서 외우실 수 있는 프로그램입니다.")
    print("="*40)
    t.sleep(1)
    print()

# 영어가 키인 기본 단어장 불러오기
def loadEngVoca(filename):
    engWordDict = {}
    with open(filename, 'r', encoding='utf-8-sig') as csvfile:
        reader = csv.reader(csvfile)
        
        for row in reader:
            if len(row) == 2: # 줄에 영어, 뜻 2개의 항목이 있을때만 저장
                eng, kor = row
                engWordDict[eng.strip()] = kor.strip()
    return engWordDict

# 뜻이 키인 기본 단어장 불러오기
def loadKorVoca(engDict):
    # 딕셔너리 컴프리헨션
    korWordDict = {k: e for e, k in engDict.items()}
    return korWordDict

# 메인 선택
def mainChoice():
    print("다음 중 무엇을 하시겠습니까?")
    print("1. 단어 시험 / 2. 단어 추가")
    return int(input())

# 출제 방식 선택
def modeChoice():
    print("\n출제 방식을 선택해주십시오.")
    print("1. 영어보고 뜻쓰기 / 2. 뜻보고 영어쓰기")
    return int(input())

# 단어 종류 선택
def kindChoice():
    print("\n단어 종류를 선택해주십시오.")
    print("1. 기본 제공 단어 (기말시험범위) / 2. 개인 단어장 단어")
    return int(input())

# 단어 시험
def wordTest(dic, filename):
    right = 0 # 맞은 개수
    textCount = 0 # 입력한 글자 수

    # csv 파일 안 단어 수 파악
    with open(filename, 'r', encoding='utf-8-sig') as f:
        word_count = sum(1 for row in csv.reader(f) if len(row) == 2)

    print(f'\n몇 개의 단어로 진행하시겠습니까? 현재 단어장 단어 개수 : {word_count}')
    while True:
        number = int(input("숫자로 입력하여 주십시오 : ")) # 시험 칠 단어 개수
        if number > word_count:
            print("입력하신 단어 수가 단어장의 단어 수보다 많습니다. 다시 입력해 주세요.")
        elif number <= 0:
            print("1 이상의 숫자를 입력해 주세요.")
        else:
            break

    input("\n[영컴타자연습] 준비되면 엔터")
    start = t.time() # 시작 시간
    for _ in range(number):
        keys = list(dic.keys())
        question = r.choice(keys) # 문제
        print(question)
        answer = input("답: ") # 답
        textCount += len(answer)
        if answer == dic[question]:
            right += 1
            print("⭕️")
            del dic[question]
        else:
            print(f'❌, 정답: {dic[question]}')

    end = t.time() # 종료 시간
    print("="*40)
    print("시험이 종료되었습니다.")
    print("-"*40)
    totalTime = end-start # 걸린시간
    print(f'걸린 시간: {totalTime:.2f}')
    print(f'맞은 개수: {right}')
    print(f'분당 타자 수: {textCount/totalTime*60:.1f}자')

# 엑셀 파일이 없으면 생성
def createExcel_ifNotExists():
    if not os.path.exists('personalWord.xlsx'):
        wb = openpyxl.Workbook()
        wb.save('personalWord.xlsx')

# 엑셀 파일 불러오기
def loadExcel():
    return openpyxl.load_workbook('personalWord.xlsx')

# 단어 추가
def inputWords(ws):
    print("영어-뜻 형식으로 입력하세요. (입력을 끝내려면 'end' 입력)")
    while True:
        user_input = input("> ").strip()
        if user_input.lower() == 'end':
            break
        elif '-' not in user_input:
            print("형식이 올바르지 않습니다. 다시 입력하십시오. (예: apple-사과)")
            continue
        english, meaning = map(str.strip, user_input.split('-', 1))
        ws.append([english, meaning])
    wb.save('personalWord.xlsx') # 엑셀 파일 저장
    print(f"\n엑셀 파일 저장 완료: 'personalWord.xlsx'")
    

# 입력했는 단어 출력
def displayWords(ws):
    print("\n입력된 단어 목록:")
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        english, meaning = row
        print(f"{idx}. {english} - {meaning}")

# 입력했던 단어 수정
def modifyWord(ws):
    while True:
        displayWords(ws)
        modification = input("\n수정할 단어가 있으면 번호를 입력하고, 없으면 엔터를 누르세요: ").strip() # 수정단어 선택
        if modification == "":
            break
        elif not modification.isdigit():
            print("숫자를 입력해주세요.")
            continue

        row_num = int(modification)
        if row_num < 1 or row_num > ws.max_row:
            print("해당 번호는 없습니다.")
            continue

        old_english = ws.cell(row=row_num, column=1).value
        old_meaning = ws.cell(row=row_num, column=2).value
        print(f"{row_num}. {old_english} - {old_meaning} 를 수정하시겠습니까?")
        check = int(input("수정: 1 / 수정 안 함: 2 > ")) # 수정 확정
        if check == 1:
            modiword = input("단어를 수정하세요. (예: cat-고양이): ").strip()
            if '-' not in modiword:
                print("형식이 올바르지 않습니다.")
                continue

            english, meaning = map(str.strip, modiword.split('-', 1))
            ws.cell(row=row_num, column=1, value=english)
            ws.cell(row=row_num, column=2, value=meaning)
            print("수정 완료")
            wb.save('personalWord.xlsx') # 엑셀 파일 저장
            print(f"\n엑셀 파일 저장 완료: 'personalWord.xlsx'")
        elif check == 2:
            continue
        else:
            print("보기에 없는 숫자입니다. 다시 선택해주십시오.")

# csv파일로 저장
def saveCsv(ws):
    with open('personalWord.csv', mode='w', newline='', encoding='utf-8-sig') as csvfile:
        writer = csv.writer(csvfile)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    print(f"\nCSV 파일 저장 완료: 'personalWord.csv'")

# 결과 후 선택
def choiceExit():
    print("프로그램을 종료할까요?")
    print("1. 프로그램 종료 / 2. 처음으로 돌아가기")
    return int(input())

# 프로그램 종료
def exitProgram():
    sys.exit(0)

# 잘못된 숫자 선택
def wrongAnswer():
    print("보기에 없는 숫자입니다. 다시 선택해주십시오.")



#main
while True:
    intro()
    main = mainChoice()

    if main == 1: # 단어 시험
        mode = modeChoice() # 영어로 칠지 한글로 칠지 선택
        kind = kindChoice() # 기본제공 단어로 칠지, 개인 단어장으로 칠지 선택

        if kind == 1: # 기본제공 단어
            enWord = loadEngVoca("word.csv") # 기본 제공 단어 단어장 불러오기
            koWord = loadKorVoca(enWord)

            if mode == 1: # 영어보고 뜻쓰기
                wordTest(enWord, "word.csv")

            elif mode == 2: # 한글보고 뜻쓰기
                wordTest(koWord, "word.csv")
            
            else: # 잘못된 입력
                wrongAnswer()
                continue
                
        elif kind == 2: # 개인 단어장 단어
            enWord = loadEngVoca("personalWord.csv") # 개인 단어장 불러오기
            koWord = loadKorVoca(enWord)

            if mode == 1: # 영어보고 뜻쓰기
                wordTest(enWord, "personalWord.csv")

            elif mode == 2: # 한글보고 뜻쓰기
                wordTest(koWord, "personalWord.csv")

            else: # 잘못된 입력
                wrongAnswer()
                continue
        
        else: # 잘못된 입력
            wrongAnswer()
            continue

        exit = choiceExit() # 종료할지 처음으로 돌아갈지 선택

        if exit == 1:
            exitProgram()
        elif exit == 2:
            continue
        else:
            wrongAnswer()
            
        
    elif main == 2: # 단어추가
        createExcel_ifNotExists()
        wb = loadExcel()
        ws = wb.active  # 기본 워크시트 불러오기
        inputWords(ws)
        modifyWord(ws)
        saveCsv(ws)

        exit = choiceExit() # 종료할지 처음으로 돌아갈지 선택

        if exit == 1:
            exitProgram()
        elif exit == 2:
            continue
        else:
            wrongAnswer()
    
    else: # 잘못된 입력
           wrongAnswer()